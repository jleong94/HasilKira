# -*- coding: utf-8 -*-
"""
LHDN Malaysia Individual Income Tax Worksheet - trilingual (ZH/BM/EN)
Reusable yearly. Supports: employment only / employment + business / other income.
Built with openpyxl.  Output: ./LHDN report/LHDN_Tax_Worksheet_<timestamp>.xlsx
"""
import os, sys, datetime
try:
    sys.stdout.reconfigure(encoding="utf-8")   # so CJK prints fine in a Windows console
except Exception:
    pass
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

# ---------- styles ----------
F_TITLE   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
F_HEAD    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_SECTION = Font(name="Calibri", size=12, bold=True, color="1F3864")
F_LABEL   = Font(name="Calibri", size=11, color="000000")
F_BOLD    = Font(name="Calibri", size=11, bold=True, color="000000")
F_NOTE    = Font(name="Calibri", size=9, italic=True, color="7F6000")
F_SMALL   = Font(name="Calibri", size=9, color="404040")

FILL_TITLE   = PatternFill("solid", fgColor="305496")
FILL_HEAD    = PatternFill("solid", fgColor="4472C4")
FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")
FILL_INPUT   = PatternFill("solid", fgColor="FFF2CC")  # yellow = user input
FILL_CALC    = PatternFill("solid", fgColor="E7E6E6")  # grey  = auto
FILL_TOTAL   = PatternFill("solid", fgColor="C6E0B4")  # green = totals
FILL_NOTE    = PatternFill("solid", fgColor="FFF9E6")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0.00'
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")

# ---------- sheet names ----------
S_COVER  = "封面 Cover"
S_PROF   = "个人资料 Profile"
S_EMP    = "就业收入 Employment"
S_BIZ    = "棕油业务 Plantation"
S_LOG    = "棕油记录 Log"
S_CA     = "免税额表 Allowance"
S_OTHER  = "其它收入 Other"
S_REL    = "扣税减免 Reliefs"
S_TAX    = "税务计算 Tax"
S_CMP    = "评估对比 Compare"
S_MY     = "多年汇总 Summary"
S_DOC    = "单据清单 Checklist"
S_LISTS  = "Lists"
CA_TOTAL_CELL = "F25"  # grand-total allowance cell on the Allowance sheet (fixed layout)

def q(sheet, ref):
    return f"'{sheet}'!{ref}"

# ---------- trilingual labels (key, zh, bm, en) ----------
LABELS = [
("title","马来西亚个人所得税计算表 (LHDN)","Lembaran Kerja Cukai Pendapatan Individu (LHDN)","Malaysia Individual Income Tax Worksheet (LHDN)"),
("lang_select","选择语言 / Pilih Bahasa / Language","Pilih Bahasa","Select Language"),
("ya","课税年 (YA)","Tahun Taksiran (YA)","Year of Assessment (YA)"),
("instructions","使用说明","Arahan","Instructions"),
("ins1","1. 在右上角选择语言（中／马／英），所有标签会自动切换。","1. Pilih bahasa di penjuru atas; semua label bertukar automatik.","1. Pick a language at the top; all labels switch automatically."),
("ins2","2. 只填写黄色格子；灰色格子是自动计算，请勿修改。","2. Isi sel KUNING sahaja; sel KELABU dikira automatik.","2. Fill YELLOW cells only; GREY cells auto-calculate."),
("ins3","3. 依序填写各分页，最后看『税务计算』得出应缴/可退税。","3. Isi setiap helaian; lihat 'Tax' untuk baki cukai.","3. Complete each sheet; see 'Tax' for the balance."),
("ins4","4. 每年可重复使用：另存新档，更新 YA 年份与数字即可。","4. Boleh guna setiap tahun: simpan salinan baru & kemas kini.","4. Reusable yearly: save a copy & update figures."),
("legend_in","黄色 = 您填写","Kuning = input anda","Yellow = your input"),
("legend_calc","灰色 = 自动计算","Kelabu = automatik","Grey = auto-calculated"),
("disclaimer","免责声明：本表仅供参考与记录，数字以 LHDN 官方与持牌税务师意见为准。YA2025（2026年报税）。","Penafian: Untuk rujukan sahaja; sahkan dengan LHDN / ejen cukai berlesen.","Disclaimer: For reference only; confirm with LHDN / a licensed tax agent."),
# profile
("profile","个人资料","Maklumat Peribadi","Personal Particulars"),
("name","姓名","Nama","Name"),
("ic_no","身份证号码","No. Kad Pengenalan","IC Number"),
("tin","税务识别号 (TIN / NPC)","No. Pengenalan Cukai (TIN)","Tax Identification No. (TIN)"),
("efile_no","报税户口 (No. e-Filing)","No. e-Filing","e-Filing No."),
("spouse_name","配偶姓名","Nama Pasangan","Spouse Name"),
("spouse_tin","配偶 TIN","TIN Pasangan","Spouse TIN"),
("assess_type","评估方式","Jenis Taksiran","Assessment Type"),
("res_status","居民身份","Status Pemastautin","Residency Status"),
("num_child","符合资格孩子人数","Bilangan Anak Layak","No. of Qualifying Children"),
("form_type","报税表格","Borang Cukai","Tax Form"),
("note_form","⚠ 有业务收入（如棕油）须用 Borang B；只有受雇收入用 Borang BE。","⚠ Pendapatan perniagaan (cth. sawit) → Borang B; gaji sahaja → Borang BE.","⚠ Business income (e.g. palm oil) → Form B; employment-only → Form BE."),
("note_tin","⚠ TIN 现为强制项目；e-Invoice／某些交易需填自己及对方（雇主／买家）的 TIN。","⚠ TIN kini wajib; e-Invois memerlukan TIN anda & pihak lain.","⚠ TIN is now mandatory; e-Invoice needs your TIN and the other party's."),
# employment
("emp_income","就业收入 (Statutory Employment Income)","Pendapatan Penggajian","Employment Income"),
("gross_salary","薪金 / 工资","Gaji","Salary / Wages"),
("bonus","花红 / 奖金","Bonus","Bonus"),
("allowance","津贴 / 佣金 / 加班","Elaun / Komisen / OT","Allowances / Commission / OT"),
("bik","实物福利 (BIK / VOLA)","Manfaat Berupa Barangan (BIK/VOLA)","Benefits-in-Kind (BIK/VOLA)"),
("gross_emp","就业收入总额","Jumlah Pendapatan Penggajian","Total Employment Income"),
("emp_contrib","供款与已扣税（供减免 / 抵扣用）","Caruman & Cukai Dipotong","Contributions & Tax Deducted (for reliefs / offset)"),
("epf_emp","雇员公积金供款 (EPF)","Caruman KWSP Pekerja","EPF (employee)"),
("socso_emp","SOCSO + EIS 供款","Caruman PERKESO + SIP","SOCSO + EIS"),
("pcb","已扣每月税 (PCB / MTD)","PCB / MTD dipotong","MTD / PCB deducted"),
("ea_note","数据来源：雇主的 EA 表格 (Borang EA / C.P.8A)。","Sumber: Borang EA majikan.","Source: employer's EA Form (C.P.8A)."),
# plantation business
("biz_income","棕油业务 (Statutory Business Income)","Pendapatan Perniagaan Sawit","Palm Oil Business Income"),
("biz_note","棕油属业务收入（第4(a)条）。须报 Borang B；可扣业务开销与资本／农业免税额。","Sawit = pendapatan perniagaan (Sek. 4(a)). Borang B; tolak perbelanjaan & elaun.","Palm oil = business income (Sec 4(a)). Use Form B; deduct expenses & allowances."),
("ffb_sales","鲜果串 (FFB) 销售总额","Jualan Buah Tandan Segar (BTS)","Fresh Fruit Bunch (FFB) sales"),
("biz_other_inc","其它业务收入","Pendapatan perniagaan lain","Other business income"),
("gross_biz","业务总收入","Jualan Kasar","Gross Business Revenue"),
("exp_header","可扣业务开销 (Allowable Expenses)","Perbelanjaan Dibenarkan","Allowable Expenses"),
("exp_fert","肥料","Baja","Fertiliser"),
("exp_pest","农药 / 除草剂","Racun / Racun rumpai","Pesticide / Weedkiller"),
("exp_labour","工钱 / 采割工 / 工人","Upah / Penuai / Pekerja","Labour / Harvesting"),
("exp_transp","运输 / 油费","Pengangkutan / Minyak","Transport / Fuel"),
("exp_quit","地税 (Cukai tanah) / 门牌税","Cukai tanah / pintu","Quit rent / Assessment"),
("exp_maint","维修 / 工具 / 机械","Penyelenggaraan / peralatan","Maintenance / Tools"),
("exp_misc","其它开销","Perbelanjaan lain","Other expenses"),
("total_exp","开销总额","Jumlah Perbelanjaan","Total Expenses"),
("cap_allow","资本免税额 / 农业免税额","Elaun Modal / Pertanian","Capital / Agriculture Allowance"),
("net_biz","业务净利 / (亏损)","Untung / (Rugi) Bersih","Net Business Profit / (Loss)"),
("rec_note","须保留所有单据与账目至少 7 年。营业额达门槛者须发 e-Invoice。","Simpan resit & rekod 7 tahun. e-Invois jika cukup ambang.","Keep receipts & records for 7 years. e-Invoice if above threshold."),
("ffb_link_note","↑ FFB 销售与其它业务收入已从『棕油记录 Log』自动汇总（灰色）。请在 Log 分页逐月填写。","↑ Jualan BTS ditarik automatik dari helaian Log. Isi data bulanan di Log.","↑ FFB sales auto-pulled from the Log sheet. Enter monthly data in Log."),
# monthly log
("log_title","棕油月度记录 (Monthly FFB Log)","Log Bulanan Sawit","Monthly Palm Oil (FFB) Log"),
("log_note","逐月填写黄色格。FFB 收入＝重量×单价，会自动汇总到『棕油业务』分页。买家与 e-Invoice 编号供记录／合规用。","Isi sel kuning setiap bulan. Pendapatan = berat×harga, ditarik ke helaian Sawit.","Fill yellow cells monthly. Income = weight×price, flows to the Plantation sheet."),
("log_month","月份","Bulan","Month"),
("log_weight","FFB 重量 (吨)","Berat BTS (tan)","FFB Weight (tonne)"),
("log_price","单价 (RM/吨)","Harga (RM/tan)","Price (RM/tonne)"),
("log_income","FFB 收入 (RM)","Pendapatan BTS (RM)","FFB Income (RM)"),
("log_otherinc","其它业务收入 (RM)","Pendapatan lain (RM)","Other biz income (RM)"),
("log_buyer","买家 / 油厂","Pembeli / Kilang","Buyer / Mill"),
("log_einv","e-Invoice 编号 / 磅单","No. e-Invois / Tiket","e-Invoice No. / Ticket"),
("total","全年总额","Jumlah Setahun","Annual Total"),
# ---- Allowance schedule ----
("ca_title","资本与农业免税额表 (Capital & Agriculture Allowance)","Jadual Elaun Modal & Pertanian","Capital & Agriculture Allowance Schedule"),
("ca_note","本表总额自动带入『棕油业务』的免税额栏。率(%)可改。若只有一个总数，可在自订行填:合格开销=该数、率=100%。","Jumlah ditarik ke helaian Sawit. Kadar boleh diubah.","Total flows to the Plantation sheet. Rates are editable."),
("ca_ag_head","农业免税额 (Agriculture Allowance) — 附表3","Elaun Pertanian — Jadual 3","Agriculture Allowance — Schedule 3"),
("ca_pm_head","资本免税额 — 机械/设备 (Capital Allowance, Plant & Machinery)","Elaun Modal — Loji & Mesin","Capital Allowance — Plant & Machinery"),
("col_desc","项目说明","Keterangan","Description"),
("col_qexp","合格开销 (RM)","Perbelanjaan Layak (RM)","Qualifying Cost (RM)"),
("col_accum","累计已扣 b/f (RM)","Elaun terkumpul b/f (RM)","Accum. allowance b/f (RM)"),
("col_rate","率 (%)","Kadar (%)","Rate (%)"),
("col_alw_year","本年免税额 (RM)","Elaun tahun ini (RM)","Allowance this year (RM)"),
("col_residual","余额 c/f (RM)","Baki c/f (RM)","Residual c/f (RM)"),
("ca_sub_ag","农业免税额小计","Subjumlah Pertanian","Agriculture subtotal"),
("ca_sub_pm","资本免税额小计","Subjumlah Modal","Capital subtotal"),
("ca_grand","免税额总计（带入棕油页）","Jumlah Elaun (ke Sawit)","Total allowance (to Plantation)"),
("ag_clear","开芭 / 整地","Membersih / menyedia tanah","Land clearing / preparation"),
("ag_plant","新种植 (棕苗)","Tanaman baru","New planting"),
("ag_road","园区道路 / 桥梁","Jalan / jambatan ladang","Farm roads / bridges"),
("ag_wbuild","工人宿舍 / 福利建筑","Bangunan kebajikan pekerja","Workers' living / welfare buildings"),
("ag_obuild","其它园区建筑","Bangunan ladang lain","Other farm buildings"),
("custom_row","自订项目 (自行填写)","Item sendiri","Custom item"),
("pm_vehicle","拖拉机 / 罗里 / 车辆","Traktor / lori / kenderaan","Tractor / lorry / vehicle"),
("pm_harvest","采割机 / 工具","Mesin menuai / alatan","Harvesting machine / tools"),
("pm_general","一般机械","Loji & mesin am","General plant & machinery"),
("pm_equip","设备 / 家具","Peralatan / perabot","Equipment / furniture"),
("pm_small","小额资产 (≤RM2,000)","Aset nilai kecil (≤RM2,000)","Small-value asset (≤RM2,000)"),
("ca_rate_note","参考率:农业—开芭/种植/道路50%、工人楼20%、其它楼10%；机械—首年=初始20%+每年(20/14/10%)，次年起只每年率。累计不超过成本。详情询税务师。","Kadar rujukan; tahun pertama = IA20% + AA. Rujuk ejen cukai.","Reference rates; first year = IA 20% + AA. Confirm with a tax agent."),
# ---- Separate vs Joint comparison ----
("cmp_title","单独 vs 联合评估对比","Banding Taksiran Berasingan vs Bersama","Separate vs Joint Assessment Comparison"),
("cmp_your_inc","你的总收入","Pendapatan anda","Your total income"),
("cmp_your_rel","你的减免（不含配偶减免）","Pelepasan anda (tanpa pelepasan pasangan)","Your reliefs (excl. spouse relief)"),
("cmp_sp_inc","配偶总收入","Pendapatan pasangan","Spouse total income"),
("cmp_sp_rel","配偶减免（单独评估时用，默认个人 RM9,000）","Pelepasan pasangan (berasingan)","Spouse reliefs (separate; default RM9,000)"),
("cmp_sep_head","方案 A：单独评估（各报各的）","Pilihan A: Taksiran Berasingan","Option A: Separate Assessment"),
("cmp_joint_head","方案 B：联合评估（合并由你申报 + 配偶减免RM4,000）","Pilihan B: Taksiran Bersama","Option B: Joint Assessment"),
("cmp_your_ci","你的应课税收入","Pendapatan bercukai anda","Your chargeable income"),
("cmp_your_tax","你的税款","Cukai anda","Your tax"),
("cmp_sp_ci","配偶应课税收入","Pendapatan bercukai pasangan","Spouse chargeable income"),
("cmp_sp_tax","配偶税款","Cukai pasangan","Spouse tax"),
("cmp_sep_total","单独评估合计税款","Jumlah cukai (berasingan)","Total tax — Separate"),
("cmp_joint_ci","联合应课税收入","Pendapatan bercukai (bersama)","Joint chargeable income"),
("cmp_joint_tax","联合评估税款","Cukai (bersama)","Total tax — Joint"),
("cmp_recommend","建议（较省税）","Cadangan (lebih jimat)","Recommended (lower tax)"),
("cmp_saving","可省税额","Penjimatan","Tax saving"),
("cmp_assume","假设:双方皆居民、用累进税率。联合评估时配偶『本身的』减免不能用，只能加 RM4,000 配偶减免。仅供比较，实际以 LHDN 为准。","Andaian: kedua-dua pemastautin. Bersama = hilang pelepasan sendiri pasangan kecuali RM4,000.","Assumes both resident. Joint forfeits spouse's own reliefs except the RM4,000."),
# ---- Multi-year summary ----
("my_title","多年度汇总 (Multi-Year Summary)","Ringkasan Pelbagai Tahun","Multi-Year Summary"),
("my_note","『本年(自动)』栏自动取自『税务计算』。每年报税后，把该栏数字抄到对应年份栏(黄色)长期保存。","Lajur 'auto' dari helaian Tax. Salin ke lajur tahun setiap tahun.","The 'auto' column reads the Tax sheet. Copy it into a year column each year."),
("my_year_hdr","课税年 (YA)","Tahun Taksiran","Year of Assessment"),
("my_thisyear","本年 (自动)","Tahun ini (auto)","This year (auto)"),
("my_rebates","回扣 (RM)","Rebat (RM)","Rebates (RM)"),
("my_paid","已缴 PCB + CP500 (RM)","Dibayar PCB + CP500","Paid PCB + CP500 (RM)"),
# ---- Documents checklist ----
("doc_title","单据 / 文件保存清单 (Documents to Keep)","Senarai Simpanan Dokumen","Documents to Keep — Checklist"),
("doc_note","法令要求保留所有单据与记录至少 7 年。每年备齐后在 ✓ 栏打勾。","Wajib simpan semua resit & rekod sekurang-kurangnya 7 tahun.","Law requires keeping all receipts & records for at least 7 years."),
("doc_chk","✓","✓","✓"),
("doc_item","文件 / 单据","Dokumen","Document"),
("doc_remark","备注","Catatan","Notes"),
("d_ea","雇主 EA 表格 (Borang EA / C.P.8A)","Borang EA majikan","Employer's EA Form (C.P.8A)"),
("d_ffb","FFB 磅单 / 销售单 / 油厂 e-Invoice","Tiket BTS / e-Invois kilang","FFB weighbridge tickets / mill e-Invoice"),
("d_exp","业务开销收据(肥料/农药/工钱/运输/维修)","Resit perbelanjaan ladang","Plantation expense receipts"),
("d_quit","地税 (Cukai tanah) / 门牌税单","Bil cukai tanah / pintu","Quit rent / assessment bills"),
("d_capital","机械 / 资产购买发票(资本免税额)","Invois aset / mesin","Asset / machinery purchase invoices"),
("d_cp500","CP500 / CP502 通知 & 缴款收据","Notis CP500/CP502 & resit bayaran","CP500/CP502 notices & payment receipts"),
("d_pcb","PCB / MTD 扣税记录 (薪资单)","Rekod PCB (slip gaji)","MTD/PCB records (payslips)"),
("d_medical","医疗收据(本人/配偶/孩子/父母)","Resit perubatan","Medical receipts (self/spouse/child/parents)"),
("d_lifestyle","生活方式 / 运动 / 电脑 / 书籍收据","Resit gaya hidup / sukan","Lifestyle / sports / device / book receipts"),
("d_edu","教育费 / 课程收据","Resit yuran pendidikan","Education / course fee receipts"),
("d_insurance","人寿 / 医疗 / 教育保险单 & EPF/PRS 报表","Penyata insurans & KWSP/PRS","Insurance statements & EPF/PRS statements"),
("d_sspn","SSPN 存款证明","Penyata SSPN","SSPN deposit statements"),
("d_childcare","托儿所 / 幼儿园收据","Resit taska / tadika","Childcare / kindergarten receipts"),
("d_donation","核准捐款收据","Resit derma diluluskan","Approved donation receipts"),
("d_zakat","天课 / Zakat / Fitrah 收据","Resit zakat / fitrah","Zakat / Fitrah receipts"),
("d_income","银行月结单 / 股息凭单 / 租约","Penyata bank / dividen / sewa","Bank statements / dividend vouchers / tenancy"),
("d_prior","上一年报税表 & e-Filing 收条","Borang & resit e-Filing tahun lalu","Prior-year return & e-Filing acknowledgement"),
("d_tin","TIN / IC 副本(自己 + 配偶)","Salinan TIN / IC (diri & pasangan)","TIN / IC copies (self & spouse)"),
# other income
("other_income","其它收入 (Other Income)","Pendapatan Lain","Other Income"),
("rental","租金收入（净额）","Pendapatan Sewa (bersih)","Rental income (net)"),
("interest","利息（须课税部分）","Faedah (bercukai)","Interest (taxable)"),
("dividend","股息（须课税部分）","Dividen (bercukai)","Dividends (taxable)"),
("royalty","版税 / 佣金 / 兼职","Royalti / Komisen","Royalty / Commission"),
("other_misc","其它","Lain-lain","Others"),
("total_other","其它收入总额","Jumlah Pendapatan Lain","Total Other Income"),
("div_note","注：自 YA2025 起，个人股息年收入超过 RM100,000 部分另征 2% 税。","Nota: Mulai YA2025, dividen >RM100,000 dikenakan cukai 2%.","Note: From YA2025, individual dividends above RM100,000 taxed at 2%."),
# reliefs
("reliefs","扣税减免 (Tax Reliefs) — YA 2025","Pelepasan Cukai — YA 2025","Tax Reliefs — YA 2025"),
("col_item","项目","Item","Item"),
("col_max","单项顶限 (RM)","Had (RM)","Max (RM)"),
("col_qty","数量","Bil.","Qty"),
("col_cap","实际顶限 (RM)","Had Sebenar","Eff. Cap"),
("col_claim","您申报 (RM)","Tuntutan (RM)","Your Claim (RM)"),
("col_allow","可扣 (RM)","Dibenarkan (RM)","Allowed (RM)"),
("total_relief","减免总额","Jumlah Pelepasan","Total Reliefs"),
("r_individual","个人与受扶养亲属（自动）","Individu & saudara tanggungan","Individual & dependent relatives"),
("r_disabled_self","残疾人士本人（额外）","Individu OKU (tambahan)","Disabled individual (additional)"),
("r_spouse","配偶 / 赡养费","Suami/Isteri / Nafkah","Spouse / Alimony"),
("r_disabled_spouse","残疾配偶","Pasangan OKU","Disabled spouse"),
("r_child_u18","孩子（未满18岁）每名","Anak < 18 (setiap)","Child under 18 (each)"),
("r_child_tert","孩子（18岁+ 高等教育）每名","Anak 18+ pengajian tinggi (setiap)","Child 18+ in tertiary (each)"),
("r_child_dis","残疾未婚孩子 每名","Anak OKU belum kahwin (setiap)","Disabled unmarried child (each)"),
("r_edu_self","自我进修教育费（含技能课程上限RM2,000）","Yuran pendidikan sendiri","Self education fees (skills sublimit RM2,000)"),
("r_med_serious","医疗：重病/生育/疫苗/牙科/体检/心理（本人/配偶/孩子）","Perubatan diri/pasangan/anak","Medical: serious illness/fertility/vaccine/dental/checkup"),
("r_med_parents","父母医疗 / 护理 / 体检","Rawatan ibu bapa","Parents medical / care / checkup"),
("r_equipment","残疾辅助器材（本人/配偶/孩子/父母）","Peralatan sokongan OKU","Disabled supporting equipment"),
("r_lifestyle","生活方式：书籍/电脑/手机/网络/课程","Gaya hidup","Lifestyle: books/PC/phone/internet"),
("r_sports","运动器材 / 健身房 / 比赛 / 培训","Sukan: peralatan/gim","Sports equipment/gym/training"),
("r_ev","电动车充电设备 / 厨余处理机","Pengecas EV / mesin kompos","EV charging / food-waste composting"),
("r_breastfeed","母乳喂养器材（2年一次）","Peralatan penyusuan (2 thn sekali)","Breastfeeding equipment (every 2 yrs)"),
("r_childcare","托儿所 / 幼儿园费（孩子≤6岁）","Yuran taska / tadika (≤6 thn)","Childcare / kindergarten (child ≤6)"),
("r_sspn","SSPN 教育储蓄（净存款）","Simpanan SSPN (bersih)","SSPN net savings"),
("r_epf","雇员公积金 (EPF)","KWSP","EPF"),
("r_life","人寿保险 / Takaful","Insurans hayat / Takaful","Life insurance / Takaful"),
("r_prs","私人退休计划 (PRS) / 递延年金","PRS / Anuiti tertunda","PRS / Deferred annuity"),
("r_eduins","教育与医疗保险","Insurans pendidikan & perubatan","Education & medical insurance"),
("r_socso","SOCSO / PERKESO 供款","Caruman PERKESO","SOCSO contribution"),
("r_house","首间房屋贷款利息（条件适用）","Faedah pinjaman rumah pertama","First-home loan interest (conditions)"),
("rel_note","部分项目有附属顶限与条件（如体检RM1,000、孩子学习障碍RM6,000）。请保留单据。","Sub-had & syarat terpakai. Simpan resit.","Sub-limits & conditions apply. Keep receipts."),
# tax computation
("tax_comp","税务计算 (Tax Computation)","Pengiraan Cukai","Tax Computation"),
("si_emp","就业法定收入","Pendapatan berkanun penggajian","Statutory income — employment"),
("si_biz","业务法定收入（棕油）","Pendapatan berkanun perniagaan","Statutory income — business"),
("si_other","其它收入","Pendapatan lain","Other income"),
("si_spouse","配偶总收入（仅『联合评估』时计入）","Pendapatan pasangan (taksiran bersama sahaja)","Spouse total income (joint assessment only)"),
("aggregate","总收入 (Aggregate Income)","Pendapatan Agregat","Aggregate Income"),
("donations","减：核准捐款（≤总收入10%）","Tolak: Derma diluluskan (≤10%)","Less: Approved donations (≤10%)"),
("total_inc","总入息 (Total Income)","Jumlah Pendapatan","Total Income"),
("less_relief","减：减免总额","Tolak: Jumlah pelepasan","Less: Total reliefs"),
("chargeable","应课税收入 (Chargeable Income)","Pendapatan Bercukai","Chargeable Income"),
("tax_charged","应缴税款（按税率表）","Cukai (jadual kadar)","Tax charged (per rate table)"),
("reb_400","减：个人回扣（应税≤RM35,000 → RM400）","Tolak: Rebat individu","Less: Individual rebate (≤RM35,000)"),
("reb_spouse","减：配偶回扣（联合评估，符合条件 RM400）","Tolak: Rebat pasangan","Less: Spouse rebate (joint, if eligible)"),
("reb_zakat","减：天课 / Zakat / Fitrah","Tolak: Zakat / Fitrah","Less: Zakat / Fitrah"),
("tax_after","回扣后应缴税","Cukai selepas rebat","Tax after rebates"),
("less_pcb","减：已扣 PCB / MTD","Tolak: PCB / MTD","Less: MTD / PCB paid"),
("less_cp500","减：CP500 分期付款","Tolak: Ansuran CP500","Less: CP500 installments"),
("balance","应补缴 / (可退税)","Baki Perlu Bayar / (Lebihan)","Balance Payable / (Refund)"),
("cp500_note","非受雇收入（如棕油）LHDN 或会发 CP500 估税分期；记得抵扣已缴部分。","Pendapatan bukan gaji: CP500 boleh dikeluarkan.","Non-employment income may get CP500 installment notices."),
("res_note","税率随『个人资料』的居民身份自动切换：居民＝累进税率＋减免；非居民＝单一 30%、无减免与回扣。","Kadar ikut status pemastautin: Pemastautin = berperingkat + pelepasan; Bukan = 30% rata.","Rate follows residency in Profile: Resident = graduated + reliefs; Non-resident = flat 30%, no reliefs/rebates."),
("joint_note","『联合评估』：在上方填配偶总收入会自动并入；记得在减免页申报配偶减免 RM4,000。","Taksiran bersama: isi pendapatan pasangan di atas; tuntut pelepasan pasangan RM4,000.","Joint assessment: enter spouse income above; claim the RM4,000 spouse relief on the Reliefs sheet."),
]

LABEL_INDEX = {k:i for i,(k,_,_,_) in enumerate(LABELS)}

# ---------- workbook ----------
wb = openpyxl.Workbook()
ws_cover  = wb.active; ws_cover.title = S_COVER
ws_prof   = wb.create_sheet(S_PROF)
ws_emp    = wb.create_sheet(S_EMP)
ws_biz    = wb.create_sheet(S_BIZ)
ws_log    = wb.create_sheet(S_LOG)
ws_ca     = wb.create_sheet(S_CA)
ws_other  = wb.create_sheet(S_OTHER)
ws_rel    = wb.create_sheet(S_REL)
ws_tax    = wb.create_sheet(S_TAX)
ws_cmp    = wb.create_sheet(S_CMP)
ws_my     = wb.create_sheet(S_MY)
ws_doc    = wb.create_sheet(S_DOC)
ws_lists  = wb.create_sheet(S_LISTS)

def L(key):
    """trilingual label via VLOOKUP on language switch"""
    return f'=IFERROR(VLOOKUP("{key}",LblTbl,LangCol,FALSE),"{key}")'

def put_label(ws, cell, key, font=F_LABEL, fill=None, align=LEFT):
    c = ws[cell]; c.value = L(key); c.font = font; c.alignment = align
    if fill: c.fill = fill
    return c

def put_input(ws, cell, value=0, fmt=MONEY):
    c = ws[cell]; c.value = value; c.fill = FILL_INPUT; c.border = BORDER
    c.number_format = fmt; c.alignment = RIGHT; c.protection = Protection(locked=False)
    return c

def put_calc(ws, cell, formula, fmt=MONEY, fill=FILL_CALC, font=F_BOLD):
    c = ws[cell]; c.value = formula; c.fill = fill; c.border = BORDER
    c.number_format = fmt; c.alignment = RIGHT; c.font = font
    return c

def section(ws, row, key, span=6):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1+span)
    c = ws.cell(row=row, column=2, value=L(key)); c.font = F_SECTION; c.fill = FILL_SECTION
    c.alignment = LEFT
    ws.row_dimensions[row].height = 22

def title_bar(ws, span=6):
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+span)
    c = ws.cell(row=1, column=2, value=L("title")); c.font = F_TITLE; c.fill = FILL_TITLE
    c.alignment = CENTER
    ws.row_dimensions[1].height = 34

def note(ws, row, key, span=6):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1+span)
    c = ws.cell(row=row, column=2, value=L(key)); c.font = F_NOTE; c.fill = FILL_NOTE
    c.alignment = LEFT
    ws.row_dimensions[row].height = 30

# ===================== LISTS SHEET =====================
ws_lists["A1"]="Key"; ws_lists["B1"]="ZH"; ws_lists["C1"]="BM"; ws_lists["D1"]="EN"
for i,(k,zh,bm,en) in enumerate(LABELS):
    r = i+2
    ws_lists.cell(row=r, column=1, value=k)
    ws_lists.cell(row=r, column=2, value=zh)
    ws_lists.cell(row=r, column=3, value=bm)
    ws_lists.cell(row=r, column=4, value=en)
for col,w in {"A":18,"B":52,"C":52,"D":52}.items():
    ws_lists.column_dimensions[col].width = w
for cell in ["A1","B1","C1","D1"]:
    ws_lists[cell].font = F_HEAD; ws_lists[cell].fill = FILL_HEAD

# language switch helper
ws_lists["G1"]="中文"; ws_lists["G2"]="Bahasa Melayu"; ws_lists["G3"]="English"
ws_lists["H1"]=f"=MATCH({q(S_COVER,'$C$3')},$G$1:$G$3,0)"
ws_lists["H2"]="=H1+1"
ws_lists["F1"]="LangIndex"; ws_lists["F2"]="LangCol"
# residency / assessment flags (1/0)
ws_lists["F4"]="IsResident"; ws_lists["H4"]=f"=IF({q(S_PROF,'$C$12')}=$O$2,1,0)"
ws_lists["F5"]="IsJoint";    ws_lists["H5"]=f"=IF({q(S_PROF,'$C$11')}=$N$3,1,0)"

# tax bracket table YA2025 : Lower | Base | Rate
ws_lists["J1"]="Lower"; ws_lists["K1"]="Base"; ws_lists["L1"]="Rate"
bands = [
 (0,0,0.00),(5000,0,0.01),(20000,150,0.03),(35000,600,0.06),(50000,1500,0.11),
 (70000,3700,0.19),(100000,9400,0.25),(400000,84400,0.26),(600000,136400,0.28),(2000000,528400,0.30),
]
for i,(lo,ba,ra) in enumerate(bands):
    r=i+2
    ws_lists.cell(row=r,column=10,value=lo)
    ws_lists.cell(row=r,column=11,value=ba)
    ws_lists.cell(row=r,column=12,value=ra)
for c in ["J1","K1","L1"]:
    ws_lists[c].font=F_HEAD; ws_lists[c].fill=FILL_HEAD

# dropdown option lists
ws_lists["N1"]="AssessOpt"; ws_lists["O1"]="ResOpt"; ws_lists["P1"]="FormOpt"; ws_lists["Q1"]="YearOpt"
assess=["单独 Berasingan Separate","联合 Bersama Joint"]
res=["居民 Pemastautin Resident","非居民 Bukan Pemastautin Non-resident"]
forms=["BE (受雇 Employment only)","B (业务 With business income)"]
for i,v in enumerate(assess): ws_lists.cell(row=i+2,column=14,value=v)
for i,v in enumerate(res):    ws_lists.cell(row=i+2,column=15,value=v)
for i,v in enumerate(forms):  ws_lists.cell(row=i+2,column=16,value=v)
ws_lists["S1"]="ChkOpt"
for i,v in enumerate(["✓","✗","N/A"]): ws_lists.cell(row=i+2,column=19,value=v)
# Year list = auto-rolling window (current year -2 .. +3) using ordinary formulas
# (no SEQUENCE / spill -> works in every Excel version, still auto-updates each year)
ws_lists["Q2"]="=YEAR(TODAY())-2"
for i in range(3,8):
    ws_lists[f"Q{i}"]=f"=Q{i-1}+1"
for c in ["N1","O1","P1","Q1"]:
    ws_lists[c].font=F_HEAD; ws_lists[c].fill=FILL_HEAD
for col in ["G","J","K","L","N","O","P","Q"]:
    ws_lists.column_dimensions[col].width=16

# defined names
def add_name(name, ref):
    wb.defined_names.add(DefinedName(name, attr_text=ref))
add_name("LblTbl", q(S_LISTS,"$A$2:$D$"+str(len(LABELS)+1)))
add_name("LangCol", q(S_LISTS,"$H$2"))
add_name("LangOpts", q(S_LISTS,"$G$1:$G$3"))
add_name("TaxLower", q(S_LISTS,"$J$2:$J$11"))
add_name("TaxBase",  q(S_LISTS,"$K$2:$K$11"))
add_name("TaxRate",  q(S_LISTS,"$L$2:$L$11"))
add_name("AssessOpts", q(S_LISTS,"$N$2:$N$3"))
add_name("ResOpts",    q(S_LISTS,"$O$2:$O$3"))
add_name("FormOpts",   q(S_LISTS,"$P$2:$P$3"))
add_name("YearOpts",   q(S_LISTS,"$Q$2:$Q$7"))  # auto-rolling 6-year window
add_name("IsResident", q(S_LISTS,"$H$4"))
add_name("IsJoint",    q(S_LISTS,"$H$5"))
add_name("ChkOpts",    q(S_LISTS,"$S$2:$S$4"))

# ===================== COVER =====================
ws = ws_cover
ws.column_dimensions["A"].width=2
ws.column_dimensions["B"].width=40
ws.column_dimensions["C"].width=40
for col in "DEFG": ws.column_dimensions[col].width=18
title_bar(ws)
# language selector
put_label(ws,"B3","lang_select",font=F_BOLD)
csel = ws["C3"]; csel.value="中文"; csel.fill=FILL_INPUT; csel.border=BORDER; csel.alignment=CENTER; csel.font=F_BOLD
dv_lang = DataValidation(type="list", formula1="LangOpts", allow_blank=False)
ws.add_data_validation(dv_lang); dv_lang.add(csel)
# year
put_label(ws,"B4","ya",font=F_BOLD)
yc=ws["C4"]; yc.value=2025; yc.fill=FILL_INPUT; yc.border=BORDER; yc.alignment=CENTER; yc.font=F_BOLD
dv_year=DataValidation(type="list", formula1="YearOpts", allow_blank=True)
ws.add_data_validation(dv_year); dv_year.add(yc)

put_label(ws,"B6","instructions",font=F_SECTION)
for i,key in enumerate(["ins1","ins2","ins3","ins4"]):
    ws.merge_cells(start_row=7+i,start_column=2,end_row=7+i,end_column=5)
    c=ws.cell(row=7+i,column=2,value=L(key)); c.font=F_LABEL; c.alignment=LEFT
# legend
ws["B12"].value=L("legend_in"); ws["B12"].font=F_SMALL; ws["B12"].fill=FILL_INPUT; ws["B12"].border=BORDER
ws["B13"].value=L("legend_calc"); ws["B13"].font=F_SMALL; ws["B13"].fill=FILL_CALC; ws["B13"].border=BORDER
note(ws,15,"disclaimer",span=4)
ws.sheet_view.showGridLines=False

# ===================== PROFILE =====================
ws=ws_prof
ws.column_dimensions["A"].width=2
ws.column_dimensions["B"].width=34
ws.column_dimensions["C"].width=34
for col in "DEFG": ws.column_dimensions[col].width=16
title_bar(ws,span=4); section(ws,3,"profile",span=4)
rows=[("name",""),("ic_no",""),("tin",""),("efile_no",""),("spouse_name",""),
      ("spouse_tin",""),("assess_type","__assess"),("res_status","__res"),
      ("num_child",0),("form_type","__form")]
r=5
dv_assess=DataValidation(type="list",formula1="AssessOpts",allow_blank=True); ws.add_data_validation(dv_assess)
dv_res=DataValidation(type="list",formula1="ResOpts",allow_blank=True); ws.add_data_validation(dv_res)
dv_form=DataValidation(type="list",formula1="FormOpts",allow_blank=True); ws.add_data_validation(dv_form)
for key,val in rows:
    put_label(ws,f"B{r}",key,font=F_BOLD)
    c=ws[f"C{r}"]; c.fill=FILL_INPUT; c.border=BORDER; c.alignment=LEFT
    if val=="__assess": c.value=assess[0]; dv_assess.add(c)
    elif val=="__res": c.value=res[0]; dv_res.add(c)
    elif val=="__form": c.value=forms[1]; dv_form.add(c)
    elif key=="num_child": c.value=0; c.alignment=RIGHT
    else: c.value=val
    r+=1
note(ws,r+1,"note_form",span=4); note(ws,r+2,"note_tin",span=4)
ws.sheet_view.showGridLines=False

# ===================== EMPLOYMENT =====================
ws=ws_emp
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=42
for col in "CD": ws.column_dimensions[col].width=18
title_bar(ws,span=3); section(ws,3,"emp_income",span=3)
put_label(ws,"B5","gross_salary"); put_input(ws,"D5")
put_label(ws,"B6","bonus"); put_input(ws,"D6")
put_label(ws,"B7","allowance"); put_input(ws,"D7")
put_label(ws,"B8","bik"); put_input(ws,"D8")
put_label(ws,"B9","gross_emp",font=F_BOLD)
put_calc(ws,"D9","=SUM(D5:D8)",fill=FILL_TOTAL)
section(ws,11,"emp_contrib",span=3)  # subheader for contribution / deduction rows
put_label(ws,"B12","epf_emp"); put_input(ws,"D12")
put_label(ws,"B13","socso_emp"); put_input(ws,"D13")
put_label(ws,"B14","pcb"); put_input(ws,"D14")
note(ws,16,"ea_note",span=3)
ws.sheet_view.showGridLines=False

# ===================== PLANTATION =====================
ws=ws_biz
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=42
for col in "CD": ws.column_dimensions[col].width=18
title_bar(ws,span=3); section(ws,3,"biz_income",span=3)
note(ws,4,"biz_note",span=3)
put_label(ws,"B6","ffb_sales"); put_calc(ws,"D6","="+q(S_LOG,"$E$18"))
put_label(ws,"B7","biz_other_inc"); put_calc(ws,"D7","="+q(S_LOG,"$F$18"))
put_label(ws,"B8","gross_biz",font=F_BOLD); put_calc(ws,"D8","=SUM(D6:D7)",fill=FILL_TOTAL)
note(ws,9,"ffb_link_note",span=3)
section(ws,10,"exp_header",span=3)
exp_rows=["exp_fert","exp_pest","exp_labour","exp_transp","exp_quit","exp_maint","exp_misc"]
r=11
for key in exp_rows:
    put_label(ws,f"B{r}",key); put_input(ws,f"D{r}"); r+=1
put_label(ws,f"B{r}","total_exp",font=F_BOLD); put_calc(ws,f"D{r}",f"=SUM(D11:D{r-1})",fill=FILL_TOTAL)
total_exp_row=r
r+=1
put_label(ws,f"B{r}","cap_allow"); put_calc(ws,f"D{r}","="+q(S_CA,"$"+CA_TOTAL_CELL[0]+"$"+CA_TOTAL_CELL[1:]))
cap_row=r
r+=1
put_label(ws,f"B{r}","net_biz",font=F_BOLD)
put_calc(ws,f"D{r}",f"=D8-D{total_exp_row}-D{cap_row}",fill=FILL_TOTAL)
NET_BIZ_CELL=f"D{r}"
note(ws,r+2,"rec_note",span=3)
ws.sheet_view.showGridLines=False

# ===================== MONTHLY FFB LOG =====================
ws=ws_log
ws.column_dimensions["A"].width=2
widths={"B":14,"C":14,"D":14,"E":15,"F":15,"G":22,"H":22}
for col,w in widths.items(): ws.column_dimensions[col].width=w
# title
ws.merge_cells("B1:H1")
c=ws["B1"]; c.value=L("log_title"); c.font=F_TITLE; c.fill=FILL_TITLE; c.alignment=CENTER
ws.row_dimensions[1].height=30
# note
ws.merge_cells("B2:H3")
c=ws["B2"]; c.value=L("log_note"); c.font=F_NOTE; c.fill=FILL_NOTE; c.alignment=LEFT
# header row 5
hdr=["log_month","log_weight","log_price","log_income","log_otherinc","log_buyer","log_einv"]
for col,key in zip("BCDEFGH",hdr):
    cc=ws[f"{col}5"]; cc.value=L(key); cc.font=F_HEAD; cc.fill=FILL_HEAD; cc.alignment=CENTER; cc.border=BORDER
# months 6..17
for i in range(12):
    r=6+i
    m=ws[f"B{r}"]; m.value=i+1; m.alignment=CENTER; m.fill=FILL_CALC; m.border=BORDER
    for col in "CD":      # weight, price = input
        x=ws[f"{col}{r}"]; x.fill=FILL_INPUT; x.border=BORDER; x.number_format=MONEY; x.alignment=RIGHT
    inc=ws[f"E{r}"]; inc.value=f"=C{r}*D{r}"; inc.fill=FILL_CALC; inc.border=BORDER; inc.number_format=MONEY; inc.alignment=RIGHT
    fo=ws[f"F{r}"]; fo.fill=FILL_INPUT; fo.border=BORDER; fo.number_format=MONEY; fo.alignment=RIGHT  # other income input
    for col in "GH":      # buyer, e-invoice text
        x=ws[f"{col}{r}"]; x.fill=FILL_INPUT; x.border=BORDER; x.alignment=LEFT
# total row 18
tr=18
tl=ws[f"B{tr}"]; tl.value=L("total"); tl.font=F_BOLD; tl.fill=FILL_TOTAL; tl.border=BORDER; tl.alignment=LEFT
for col in "CEF":
    x=ws[f"{col}{tr}"]; x.value=f"=SUM({col}6:{col}17)"; x.font=F_BOLD; x.fill=FILL_TOTAL; x.border=BORDER; x.number_format=MONEY; x.alignment=RIGHT
ws["D18"].fill=FILL_TOTAL; ws["D18"].border=BORDER
ws.sheet_view.showGridLines=False

# ===================== OTHER INCOME =====================
ws=ws_other
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=42
for col in "CD": ws.column_dimensions[col].width=18
title_bar(ws,span=3); section(ws,3,"other_income",span=3)
oth_rows=["rental","interest","dividend","royalty","other_misc"]
r=5
for key in oth_rows:
    put_label(ws,f"B{r}",key); put_input(ws,f"D{r}"); r+=1
put_label(ws,f"B{r}","total_other",font=F_BOLD); put_calc(ws,f"D{r}",f"=SUM(D5:D{r-1})",fill=FILL_TOTAL)
TOTAL_OTHER_CELL=f"D{r}"
note(ws,r+2,"div_note",span=3)
ws.sheet_view.showGridLines=False

# ===================== RELIEFS =====================
ws=ws_rel
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=54
for col in "CDEFG": ws.column_dimensions[col].width=15
title_bar(ws,span=6); section(ws,3,"reliefs",span=6)
# header row
hr=5
for col,key in zip("BCDEFG",["col_item","col_max","col_qty","col_cap","col_claim","col_allow"]):
    c=ws[f"{col}{hr}"]; c.value=L(key); c.font=F_HEAD; c.fill=FILL_HEAD; c.alignment=CENTER; c.border=BORDER
# relief data: (key, per_unit_max, qty_editable, claim_link)
reliefs=[
 ("r_individual",9000,False,9000),
 ("r_disabled_self",7000,False,0),
 ("r_spouse",4000,False,0),
 ("r_disabled_spouse",6000,False,0),
 ("r_child_u18",2000,True,0),
 ("r_child_tert",8000,True,0),
 ("r_child_dis",8000,True,0),
 ("r_edu_self",7000,False,0),
 ("r_med_serious",10000,False,0),
 ("r_med_parents",8000,False,0),
 ("r_equipment",6000,False,0),
 ("r_lifestyle",2500,False,0),
 ("r_sports",1000,False,0),
 ("r_ev",2500,False,0),
 ("r_breastfeed",1000,False,0),
 ("r_childcare",3000,False,0),
 ("r_sspn",8000,False,0),
 ("r_epf",4000,False,"=MIN("+q(S_EMP,'$D$12')+",4000)"),
 ("r_life",3000,False,0),
 ("r_prs",3000,False,0),
 ("r_eduins",3000,False,0),
 ("r_socso",350,False,"="+q(S_EMP,'$D$13')),
 ("r_house",7000,False,0),
]
start=6
r=start
SPOUSE_RELIEF_CELL=None
for key,maxv,qty_edit,claim in reliefs:
    if key=="r_spouse": SPOUSE_RELIEF_CELL=f"G{r}"
    put_label(ws,f"B{r}",key)
    ws[f"B{r}"].border=BORDER
    # max
    cmax=ws[f"C{r}"]; cmax.value=maxv; cmax.number_format=MONEY; cmax.alignment=RIGHT; cmax.fill=FILL_CALC; cmax.border=BORDER
    # qty
    cqty=ws[f"D{r}"]; cqty.value=1; cqty.alignment=CENTER; cqty.border=BORDER
    if qty_edit: cqty.fill=FILL_INPUT
    else: cqty.fill=FILL_CALC
    # eff cap
    ccap=ws[f"E{r}"]; ccap.value=f"=C{r}*D{r}"; ccap.number_format=MONEY; ccap.alignment=RIGHT; ccap.fill=FILL_CALC; ccap.border=BORDER
    # claim (input or linked)
    cclaim=ws[f"F{r}"]
    if isinstance(claim,str):
        cclaim.value=claim; cclaim.fill=FILL_CALC
    else:
        cclaim.value=claim; cclaim.fill=FILL_INPUT
    cclaim.number_format=MONEY; cclaim.alignment=RIGHT; cclaim.border=BORDER
    # allowed
    callow=ws[f"G{r}"]; callow.value=f"=MIN(F{r},E{r})"; callow.number_format=MONEY; callow.alignment=RIGHT; callow.fill=FILL_TOTAL; callow.border=BORDER; callow.font=F_BOLD
    r+=1
end=r-1
# total
put_label(ws,f"B{r}","total_relief",font=F_BOLD); ws[f"B{r}"].fill=FILL_TOTAL; ws[f"B{r}"].border=BORDER
ws.merge_cells(f"C{r}:F{r}")
tot=ws[f"G{r}"]; tot.value=f"=SUM(G{start}:G{end})"; tot.number_format=MONEY; tot.alignment=RIGHT; tot.fill=FILL_TOTAL; tot.font=F_BOLD; tot.border=BORDER
TOTAL_RELIEF_CELL=f"G{r}"
note(ws,r+2,"rel_note",span=6)
ws.sheet_view.showGridLines=False

# ===================== TAX COMPUTATION =====================
ws=ws_tax
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=48
for col in "CD": ws.column_dimensions[col].width=20
title_bar(ws,span=3); section(ws,3,"tax_comp",span=3)
REL_TOTAL_REF="="+q(S_REL,"$"+TOTAL_RELIEF_CELL[0]+"$"+TOTAL_RELIEF_CELL[1:])
NET_BIZ_REF="="+q(S_BIZ,'$'+NET_BIZ_CELL[0]+'$'+NET_BIZ_CELL[1:])
OTHER_REF="="+q(S_OTHER,'$'+TOTAL_OTHER_CELL[0]+'$'+TOTAL_OTHER_CELL[1:])
r=5
put_label(ws,f"B{r}","si_emp"); put_calc(ws,f"D{r}","="+q(S_EMP,'$D$9')); EMP_R=r; r+=1
put_label(ws,f"B{r}","si_biz"); put_calc(ws,f"D{r}",NET_BIZ_REF); BIZ_R=r; r+=1
put_label(ws,f"B{r}","si_other"); put_calc(ws,f"D{r}",OTHER_REF); OTH_R=r; r+=1
# spouse income (only counted under joint assessment)
put_label(ws,f"B{r}","si_spouse"); put_input(ws,f"D{r}"); SP_IN_R=r; r+=1
put_label(ws,f"B{r}","aggregate",font=F_BOLD)
put_calc(ws,f"D{r}",f"=SUM(D{EMP_R}:D{OTH_R})+IF(IsJoint=1,D{SP_IN_R},0)",fill=FILL_TOTAL); AGG_R=r; r+=1
put_label(ws,f"B{r}","donations"); put_input(ws,f"D{r}"); DON_R=r; r+=1
put_label(ws,f"B{r}","total_inc",font=F_BOLD); put_calc(ws,f"D{r}",f"=MAX(0,D{AGG_R}-MIN(D{DON_R},0.1*D{AGG_R}))",fill=FILL_TOTAL); TI_R=r; r+=1
# reliefs apply to residents only
put_label(ws,f"B{r}","less_relief"); put_calc(ws,f"D{r}",f"=IF(IsResident=1,{REL_TOTAL_REF[1:]},0)"); REL_R=r; r+=1
put_label(ws,f"B{r}","chargeable",font=F_BOLD); put_calc(ws,f"D{r}",f"=MAX(0,D{TI_R}-D{REL_R})",fill=FILL_TOTAL); CI_R=r; r+=1
# tax charged: resident graduated table OR non-resident flat 30%
ci=f"D{CI_R}"
grad=(f"LOOKUP({ci},TaxLower,TaxBase)+({ci}-LOOKUP({ci},TaxLower))*LOOKUP({ci},TaxLower,TaxRate)")
tax_formula=f"=IF(IsResident=1,{grad},0.30*{ci})"
put_label(ws,f"B{r}","tax_charged",font=F_BOLD); put_calc(ws,f"D{r}",tax_formula,fill=FILL_TOTAL); TAX_R=r; r+=1
# rebates: residents only; spouse rebate needs joint assessment
put_label(ws,f"B{r}","reb_400"); put_calc(ws,f"D{r}",f"=IF(AND(IsResident=1,D{CI_R}<=35000),400,0)"); R400_R=r; r+=1
put_label(ws,f"B{r}","reb_spouse"); put_calc(ws,f"D{r}",f"=IF(AND(IsResident=1,IsJoint=1,D{CI_R}<=35000),400,0)"); RSP_R=r; r+=1
put_label(ws,f"B{r}","reb_zakat"); put_input(ws,f"D{r}"); ZAK_R=r; r+=1
put_label(ws,f"B{r}","tax_after",font=F_BOLD); put_calc(ws,f"D{r}",f"=MAX(0,D{TAX_R}-D{R400_R}-D{RSP_R}-D{ZAK_R})",fill=FILL_TOTAL); TA_R=r; r+=1
put_label(ws,f"B{r}","less_pcb"); put_calc(ws,f"D{r}","="+q(S_EMP,'$D$14')); PCB_R=r; r+=1
put_label(ws,f"B{r}","less_cp500"); put_input(ws,f"D{r}"); CP_R=r; r+=1
put_label(ws,f"B{r}","balance",font=F_BOLD)
bc=put_calc(ws,f"D{r}",f"=D{TA_R}-D{PCB_R}-D{CP_R}",fill=FILL_TOTAL); bc.font=Font(bold=True,size=12,color="9C0006"); BAL_R=r
note(ws,r+2,"res_note",span=3)
note(ws,r+3,"joint_note",span=3)
note(ws,r+4,"cp500_note",span=3)
ws.sheet_view.showGridLines=False

# ===================== ALLOWANCE SCHEDULE =====================
ws=ws_ca
ws.column_dimensions["A"].width=2
for col,w in {"B":34,"C":16,"D":18,"E":10,"F":16,"G":16}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("B1:G1")
c=ws["B1"]; c.value=L("ca_title"); c.font=F_TITLE; c.fill=FILL_TITLE; c.alignment=CENTER; ws.row_dimensions[1].height=30
ws.merge_cells("B2:G3")
c=ws["B2"]; c.value=L("ca_note"); c.font=F_NOTE; c.fill=FILL_NOTE; c.alignment=LEFT
def ca_headers(hr):
    for col,key in zip("BCDEFG",["col_desc","col_qexp","col_accum","col_rate","col_alw_year","col_residual"]):
        x=ws[f"{col}{hr}"]; x.value=L(key); x.font=F_HEAD; x.fill=FILL_HEAD; x.alignment=CENTER; x.border=BORDER
def ca_row(r,key,rate):
    put_label(ws,f"B{r}",key); ws[f"B{r}"].border=BORDER
    for col in "CD":
        x=ws[f"{col}{r}"]; x.value=0; x.fill=FILL_INPUT; x.border=BORDER; x.number_format=MONEY; x.alignment=RIGHT
    e=ws[f"E{r}"]; e.value=rate; e.fill=FILL_INPUT; e.border=BORDER; e.number_format="0%"; e.alignment=CENTER
    f=ws[f"F{r}"]; f.value=f"=MAX(0,MIN(C{r}*E{r},C{r}-D{r}))"; f.fill=FILL_CALC; f.border=BORDER; f.number_format=MONEY; f.alignment=RIGHT; f.font=F_BOLD
    g=ws[f"G{r}"]; g.value=f"=C{r}-D{r}-F{r}"; g.fill=FILL_CALC; g.border=BORDER; g.number_format=MONEY; g.alignment=RIGHT
# Agriculture allowance
section(ws,5,"ca_ag_head",span=6); ca_headers(6)
ag_items=[("ag_clear",0.5),("ag_plant",0.5),("ag_road",0.5),("ag_wbuild",0.2),("ag_obuild",0.1),("custom_row",0)]
for i,(k,rt) in enumerate(ag_items): ca_row(7+i,k,rt)
put_label(ws,"B13","ca_sub_ag",font=F_BOLD); ws["B13"].fill=FILL_TOTAL; ws["B13"].border=BORDER
ws.merge_cells("C13:E13")
x=ws["F13"]; x.value="=SUM(F7:F12)"; x.fill=FILL_TOTAL; x.font=F_BOLD; x.border=BORDER; x.number_format=MONEY; x.alignment=RIGHT
ws["G13"].fill=FILL_TOTAL; ws["G13"].border=BORDER
# Capital allowance (plant & machinery)
section(ws,15,"ca_pm_head",span=6); ca_headers(16)
pm_items=[("pm_vehicle",0.4),("pm_harvest",0.34),("pm_general",0.34),("pm_equip",0.30),("pm_small",1.0),("custom_row",0)]
for i,(k,rt) in enumerate(pm_items): ca_row(17+i,k,rt)
put_label(ws,"B23","ca_sub_pm",font=F_BOLD); ws["B23"].fill=FILL_TOTAL; ws["B23"].border=BORDER
ws.merge_cells("C23:E23")
x=ws["F23"]; x.value="=SUM(F17:F22)"; x.fill=FILL_TOTAL; x.font=F_BOLD; x.border=BORDER; x.number_format=MONEY; x.alignment=RIGHT
ws["G23"].fill=FILL_TOTAL; ws["G23"].border=BORDER
# grand total (row 25 == CA_TOTAL_CELL)
put_label(ws,"B25","ca_grand",font=F_BOLD); ws["B25"].fill=FILL_TOTAL; ws["B25"].border=BORDER
ws.merge_cells("C25:E25")
x=ws["F25"]; x.value="=F13+F23"; x.fill=FILL_TOTAL; x.font=Font(bold=True,size=12,color="1F3864"); x.border=BORDER; x.number_format=MONEY; x.alignment=RIGHT
ws["G25"].fill=FILL_TOTAL; ws["G25"].border=BORDER
note(ws,27,"ca_rate_note",span=6)
ws.sheet_view.showGridLines=False

# ===================== SEPARATE vs JOINT COMPARISON =====================
ws=ws_cmp
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=50
for col in "CD": ws.column_dimensions[col].width=20
ws.merge_cells("B1:D1")
c=ws["B1"]; c.value=L("cmp_title"); c.font=F_TITLE; c.fill=FILL_TITLE; c.alignment=CENTER; ws.row_dimensions[1].height=30
emp_ref=q(S_EMP,'$D$9')
biz_ref=q(S_BIZ,'$'+NET_BIZ_CELL[0]+'$'+NET_BIZ_CELL[1:])
other_ref=q(S_OTHER,'$'+TOTAL_OTHER_CELL[0]+'$'+TOTAL_OTHER_CELL[1:])
trel_ref=q(S_REL,'$'+TOTAL_RELIEF_CELL[0]+'$'+TOTAL_RELIEF_CELL[1:])
sprel_ref=q(S_REL,'$'+SPOUSE_RELIEF_CELL[0]+'$'+SPOUSE_RELIEF_CELL[1:])
spinc_ref=q(S_TAX,f'$D${SP_IN_R}')
def grad(cell): return f"LOOKUP({cell},TaxLower,TaxBase)+({cell}-LOOKUP({cell},TaxLower))*LOOKUP({cell},TaxLower,TaxRate)"
put_label(ws,"B3","cmp_your_inc"); put_calc(ws,"D3",f"={emp_ref}+{biz_ref}+{other_ref}")
put_label(ws,"B4","cmp_your_rel"); put_calc(ws,"D4",f"={trel_ref}-{sprel_ref}")
put_label(ws,"B5","cmp_sp_inc"); put_calc(ws,"D5","="+spinc_ref)
put_label(ws,"B6","cmp_sp_rel"); put_input(ws,"D6",9000)
section(ws,8,"cmp_sep_head",span=3)
put_label(ws,"B9","cmp_your_ci"); put_calc(ws,"D9","=MAX(0,D3-D4)")
put_label(ws,"B10","cmp_your_tax"); put_calc(ws,"D10",f"=MAX(0,{grad('D9')}-IF(D9<=35000,400,0))")
put_label(ws,"B11","cmp_sp_ci"); put_calc(ws,"D11","=MAX(0,D5-D6)")
put_label(ws,"B12","cmp_sp_tax"); put_calc(ws,"D12",f"=MAX(0,{grad('D11')}-IF(D11<=35000,400,0))")
put_label(ws,"B13","cmp_sep_total",font=F_BOLD); put_calc(ws,"D13","=D10+D12",fill=FILL_TOTAL)
section(ws,15,"cmp_joint_head",span=3)
put_label(ws,"B16","cmp_joint_ci"); put_calc(ws,"D16","=MAX(0,(D3+D5)-(D4+4000))")
put_label(ws,"B17","cmp_joint_tax",font=F_BOLD); put_calc(ws,"D17",f"=MAX(0,{grad('D16')}-IF(D16<=35000,800,0))",fill=FILL_TOTAL)
put_label(ws,"B19","cmp_recommend",font=F_BOLD)
rc=ws["D19"]; rc.value='=IF(D13<=D17,"单独 Separate","联合 Joint")'; rc.fill=FILL_TOTAL; rc.font=F_BOLD; rc.border=BORDER; rc.alignment=CENTER
put_label(ws,"B20","cmp_saving",font=F_BOLD); put_calc(ws,"D20","=ABS(D13-D17)",fill=FILL_TOTAL)
note(ws,22,"cmp_assume",span=3)
ws.sheet_view.showGridLines=False

# ===================== MULTI-YEAR SUMMARY =====================
ws=ws_my
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=34
for col in "CDEFGHI": ws.column_dimensions[col].width=14
ws.merge_cells("B1:I1")
c=ws["B1"]; c.value=L("my_title"); c.font=F_TITLE; c.fill=FILL_TITLE; c.alignment=CENTER; ws.row_dimensions[1].height=30
ws.merge_cells("B2:I3")
c=ws["B2"]; c.value=L("my_note"); c.font=F_NOTE; c.fill=FILL_NOTE; c.alignment=LEFT
# caption marking column C as the auto (this-year) column
cap=ws["C4"]; cap.value=L("my_thisyear"); cap.font=F_SMALL; cap.alignment=CENTER
# header row 5
put_label(ws,"B5","my_year_hdr",font=F_HEAD); ws["B5"].fill=FILL_HEAD; ws["B5"].border=BORDER; ws["B5"].alignment=CENTER
ah=ws["C5"]; ah.value="="+q(S_COVER,"$C$4"); ah.font=F_HEAD; ah.fill=FILL_HEAD; ah.border=BORDER; ah.alignment=CENTER
for col in "DEFGHI":
    x=ws[f"{col}5"]; x.fill=FILL_INPUT; x.border=BORDER; x.alignment=CENTER  # year numbers (input)
metrics=[("aggregate",AGG_R),("less_relief",REL_R),("chargeable",CI_R),("tax_charged",TAX_R),
         ("my_rebates",None),("tax_after",TA_R),("my_paid",None),("balance",BAL_R)]
rr=6
for key,src in metrics:
    put_label(ws,f"B{rr}",key); ws[f"B{rr}"].border=BORDER
    cc=ws[f"C{rr}"]
    if key=="my_rebates": cc.value=f"={q(S_TAX,f'$D${R400_R}')}+{q(S_TAX,f'$D${RSP_R}')}+{q(S_TAX,f'$D${ZAK_R}')}"
    elif key=="my_paid":  cc.value=f"={q(S_TAX,f'$D${PCB_R}')}+{q(S_TAX,f'$D${CP_R}')}"
    else:                 cc.value="="+q(S_TAX,f"$D${src}")
    cc.fill=FILL_CALC; cc.border=BORDER; cc.number_format=MONEY; cc.alignment=RIGHT; cc.font=F_BOLD
    for col in "DEFGHI":
        x=ws[f"{col}{rr}"]; x.fill=FILL_INPUT; x.border=BORDER; x.number_format=MONEY; x.alignment=RIGHT
    rr+=1
ws.sheet_view.showGridLines=False

# ===================== DOCUMENTS CHECKLIST =====================
ws=ws_doc
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=6
ws.column_dimensions["C"].width=52; ws.column_dimensions["D"].width=30
ws.merge_cells("B1:D1")
c=ws["B1"]; c.value=L("doc_title"); c.font=F_TITLE; c.fill=FILL_TITLE; c.alignment=CENTER; ws.row_dimensions[1].height=30
ws.merge_cells("B2:D3")
c=ws["B2"]; c.value=L("doc_note"); c.font=F_NOTE; c.fill=FILL_NOTE; c.alignment=LEFT
for col,key in zip("BCD",["doc_chk","doc_item","doc_remark"]):
    x=ws[f"{col}5"]; x.value=L(key); x.font=F_HEAD; x.fill=FILL_HEAD; x.alignment=CENTER; x.border=BORDER
doc_items=["d_ea","d_ffb","d_exp","d_quit","d_capital","d_cp500","d_pcb","d_medical",
           "d_lifestyle","d_edu","d_insurance","d_sspn","d_childcare","d_donation",
           "d_zakat","d_income","d_prior","d_tin"]
rr=6
for key in doc_items:
    chk=ws[f"B{rr}"]; chk.fill=FILL_INPUT; chk.border=BORDER; chk.alignment=CENTER; chk.font=F_BOLD
    put_label(ws,f"C{rr}",key); ws[f"C{rr}"].border=BORDER
    rk=ws[f"D{rr}"]; rk.fill=FILL_INPUT; rk.border=BORDER; rk.alignment=LEFT
    rr+=1
# dropdown (✓ / ✗ / N/A) on the tick column
dv_chk=DataValidation(type="list", formula1="ChkOpts", allow_blank=True)
ws.add_data_validation(dv_chk); dv_chk.add(f"B6:B{rr-1}")
# colour the tick column: ✓ green, ✗ red, N/A grey
chk_range=f"B6:B{rr-1}"
ws.conditional_formatting.add(chk_range, CellIsRule(operator="equal", formula=['"✓"'],
    fill=PatternFill("solid",fgColor="C6EFCE"), font=Font(bold=True,color="006100")))
ws.conditional_formatting.add(chk_range, CellIsRule(operator="equal", formula=['"✗"'],
    fill=PatternFill("solid",fgColor="FFC7CE"), font=Font(bold=True,color="9C0006")))
ws.conditional_formatting.add(chk_range, CellIsRule(operator="equal", formula=['"N/A"'],
    fill=PatternFill("solid",fgColor="D9D9D9"), font=Font(color="808080")))
ws.sheet_view.showGridLines=False

# ---- reorder sheet tabs ----
desired=[S_COVER,S_PROF,S_EMP,S_BIZ,S_LOG,S_CA,S_OTHER,S_REL,S_TAX,S_CMP,S_MY,S_DOC,S_LISTS]
wb._sheets.sort(key=lambda s: desired.index(s.title))

# ---- LOCK formula cells: unlock ONLY yellow input cells, then protect each sheet ----
def is_input(cell):
    try:
        rgb=cell.fill.fgColor.rgb
        return isinstance(rgb,str) and rgb.endswith("FFF2CC")
    except Exception:
        return False
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for cell in row:
            if is_input(cell):
                cell.protection = Protection(locked=False)
    sh.protection.sheet = True               # enable protection (NO password -> easy to unlock)
    sh.protection.selectLockedCells = False  # locked cells still selectable/copyable
    sh.protection.selectUnlockedCells = False

# Set active to cover.
wb.active=0
# force Excel to recalculate everything on open (so VLOOKUP labels / tax all show)
wb.calculation.fullCalcOnLoad = True

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(SCRIPT_DIR, "LHDN report")
os.makedirs(OUT_DIR, exist_ok=True)
year = datetime.datetime.now().year   # current year auto-filled into the filename
out = os.path.join(OUT_DIR, f"LHDN_Tax_Worksheet_YA{year}.xlsx")
wb.save(out)
print("SAVED:", out)
